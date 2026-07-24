import os
from openpyxl import Workbook, load_workbook

class ExcelLogger:
    def __init__(self, file_path):
        self.file_path = file_path
        self.headers = [
            "Timestamp", 
            "Cycle Number", 
            "Cycle Time", 
            "Total Time", 
            "Current", 
            "Glucose Concentration"
        ]
        self._initialize_file()

    def _initialize_file(self):
        """Creates the Excel file with headers if it doesn't exist."""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Experiment Log"
            ws.append(self.headers)
            wb.save(self.file_path)

    def log_entry(self, timestamp, cycle_number, cycle_time, total_time, current, glucose_concentration):
        """Appends a new row to the Excel file securely."""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            row = [
                timestamp,
                cycle_number,
                cycle_time,
                total_time,
                current,
                glucose_concentration
            ]
            
            ws.append(row)
            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"Failed to log to Excel: {e}")
            return False
